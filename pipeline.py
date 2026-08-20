"""
신상품 업데이트 파일 생성 파이프라인

기존에 사람이 엑셀에서 손으로 하던 작업(옵션ver3.py + 자동업데이트파일생성(2026).py 두 스크립트 +
그 사이의 수작업 열복사/파일저장)을, 업로드된 두 개의 엑셀 파일(기린.xlsx, S파일)로부터
전부 인메모리로 재현합니다. 로컬 고정 경로(BASE_PATH 등) 대신 파일을 bytes로 주고받고,
모든 결과물은 {파일명: bytes} 딕셔너리로 모아 최종적으로 ZIP으로 묶습니다.

각 단계의 계산/치환 로직(정규식, 가격 계산, 상품명 가공, 옵션 파싱, 카테고리 매핑 등)은
원본 두 스크립트의 로직을 그대로 옮겨왔습니다 — 로직 자체를 바꾸지 않고, 입출력 방식만
"로컬 파일 경로" -> "업로드된 bytes / 딕셔너리"로 바꾼 것입니다.
"""

import io
import re
from collections import defaultdict
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string


# =========================================================
# 공통 유틸
# =========================================================

def safe_str(value):
    """None, nan 등을 안전하게 문자열로 변환"""
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except (TypeError, ValueError):
        pass
    return str(value)


def str_series(series):
    """DataFrame 컬럼(Series)을 안전하게 문자열 Series로 변환.
    pandas는 read_excel(dtype=str)로 읽어도, 그리고 그 뒤에 .astype(str)을 다시 호출해도
    빈 셀(NaN)은 실제로는 float(nan)으로 남아있는 경우가 있어(컬럼이 이미 'str' dtype으로
    표시되면 astype(str) 재호출이 아무 동작도 하지 않음), 나중에 '|'.join() 등에서
    "expected str instance, float found" 오류가 날 수 있습니다. fillna('')를 먼저 적용해
    빈 셀을 확실히 빈 문자열로 만든 뒤 문자열로 변환합니다."""
    return series.fillna('').astype(str)


def normalize_header(value):
    if value is None:
        return ''
    text = str(value)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r' *\n *', '\n', text)
    return text.strip()


def build_header_map(ws, header_row=2):
    headers = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col).value
        norm = normalize_header(raw)
        if norm:
            headers[norm] = col
    return headers


def get_col_index(letter):
    return column_index_from_string(letter)


def set_cell_by_header(ws, row_num, header_map, header_name, value, wrap_text=False):
    norm_name = normalize_header(header_name)
    if norm_name not in header_map:
        return
    col = header_map[norm_name]
    cell = ws.cell(row=row_num, column=col)
    cell.value = value
    if wrap_text:
        cell.alignment = Alignment(wrapText=True)


def load_wb(file_bytes, data_only=False):
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=data_only)


def wb_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def df_to_xlsx_bytes(df, sheet_name='Sheet1'):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


class PipelineError(Exception):
    pass


# =========================================================
# 0단계: 기린 -> S파일 열 복사, 곽충현/문정희 스냅샷, 예제 상태 생성
# =========================================================

def copy_columns_positional(src_wb, dst_wb, columns, header_row=1):
    """src_wb의 지정 열 값을 dst_wb의 같은 열/같은 행 위치에 그대로 복사.
    (엑셀에서 열 범위를 선택해 그대로 복사-붙여넣기 하는 것과 동일한 동작)
    행 수가 다르면 더 짧은 쪽 기준으로 복사하고 경고를 반환합니다.
    """
    ws_src = src_wb.active
    ws_dst = dst_wb.active

    n_src = ws_src.max_row
    n_dst = ws_dst.max_row
    n_rows = min(n_src, n_dst)

    warnings = []
    if n_src != n_dst:
        warnings.append(
            f"기린 파일({n_src}행)과 S파일({n_dst}행)의 행 수가 서로 달라, "
            f"앞쪽 {n_rows}행까지만 {'/'.join(columns)}열을 복사했습니다. "
            f"(엑셀에서 두 파일을 나란히 놓고 범위 복사할 때와 같은 방식으로, 행 번호 기준으로 맞춰 복사합니다)"
        )

    for col_letter in columns:
        col_idx = column_index_from_string(col_letter)
        for r in range(header_row + 1, n_rows + 1):
            ws_dst.cell(row=r, column=col_idx).value = ws_src.cell(row=r, column=col_idx).value

    return warnings


def clean_product_name_style5(name):
    """2단계 '5번시트' 그룹(오토/문정희 등)과 동일한 상품명 정리 규칙:
    '<' 이후 전부 삭제 + 리터럴 'KC' 문자열만 제거 (KC 앞부분은 그대로 유지)."""
    if not isinstance(name, str):
        return name
    name = re.sub(r'<.*', '', name)
    name = name.replace('KC', '')
    return name


def build_moon_wb(wb_s):
    """문정희 스냅샷용: wb_s(현재 상태)를 복제해 상품명 열에
    2단계 문정희 로직(5번시트 스타일 정리 + '[상시의류]' 접두어)을 적용한 새 워크북 반환."""
    moon_wb = load_wb(wb_bytes(wb_s), data_only=False)
    ws = moon_wb.active

    name_col = None
    for cell in ws[1]:
        if cell.value == '상품명':
            name_col = cell.column
            break
    if name_col is None:
        return moon_wb

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=name_col)
        if isinstance(cell.value, str):
            cell.value = '[상시의류]' + clean_product_name_style5(cell.value)

    return moon_wb


def build_example_state(kirin_bytes, sfile_bytes):
    """기린.xlsx + S파일 -> (E,F,G 복사 → 곽충현/문정희 스냅샷 → AJ 복사) -> 예제 상태 bytes

    문정희는 이 시점(E,F,G만 반영된 상태)의 상품명에 2단계 문정희 가공 규칙
    ('<' 이후 삭제 + 'KC' 제거 + '[상시의류]' 접두어)을 적용해 최종본으로 생성합니다.
    (곽충현은 아직 별도 가공 규칙이 없어 현재는 스냅샷 그대로 저장합니다.)

    반환: dict(example_bytes, kwak_bytes, moon_bytes, warnings)
    """
    warnings = []

    wb_kirin = load_wb(kirin_bytes, data_only=True)
    wb_s = load_wb(sfile_bytes, data_only=False)

    # 1) E,F,G 열 복사
    warnings += copy_columns_positional(wb_kirin, wb_s, ['E', 'F', 'G'])

    # 2) 이 시점 상태를 곽충현(스냅샷) / 문정희(가공본)로 저장
    kwak_bytes = wb_bytes(wb_s)
    moon_bytes = wb_bytes(build_moon_wb(wb_s))

    # 3) AJ 열 복사
    warnings += copy_columns_positional(wb_kirin, wb_s, ['AJ'])

    # 4) 이 상태 = 예제.xlsx
    example_bytes = wb_bytes(wb_s)

    return {
        'example_bytes': example_bytes,
        'kwak_bytes': kwak_bytes,
        'moon_bytes': moon_bytes,
        'warnings': warnings,
    }


# =========================================================
# 1단계: 옵션ver3.py 로직 (예제.xlsx -> 새 기린.xlsx / 날짜파일)
# =========================================================

def parse_option_string(option_str):
    groups = option_str.split('//')
    option_dict = {}
    has_extra_price = False
    for group in groups:
        match = re.match(r'(.+?)\{(.+?)\}', group)
        if match:
            group_name = match.group(1).strip()
            options = match.group(2).split('|')
            parsed_options = []
            for opt in options:
                opt = opt.strip()
                add_price_match = re.match(r'(.+?)\((.*?\+(\d+))\)', opt)
                if add_price_match:
                    opt_name = add_price_match.group(1)
                    extra = int(add_price_match.group(3))
                    parsed_options.append((opt_name, extra))
                    has_extra_price = True
                else:
                    parsed_options.append((opt, 0))
            option_dict[group_name] = parsed_options
    return option_dict, has_extra_price


def group_sizes_by_extra_preserve_color(option_dict):
    size_group = option_dict.get("사이즈", [])
    size_price_map = defaultdict(list)
    for size, extra in size_group:
        size_price_map[extra].append(size)
    grouped_sizes = []
    for extra_price, sizes in size_price_map.items():
        grouped_sizes.append({"extra_price": extra_price, "sizes": sizes})
    return grouped_sizes


def update_product_name(original_name, sizes, apply_update):
    if not apply_update or not sizes:
        return original_name
    first = sizes[0]
    last = sizes[-1]
    size_range = f"{first}" if first == last else f"{first}~{last}"
    updated_name = re.sub(r'\*.*?\*', f"*{size_range}*", original_name)
    if "<br>" in updated_name:
        parts = updated_name.split("<br>", 1)
        updated_name = f"{parts[0]}({size_range})<br>{parts[1]}"
    else:
        updated_name = f"{updated_name}({size_range})"
    return updated_name


def expand_rows(df):
    all_expanded = []
    errors = []

    for idx, row in df.iterrows():
        option_str = row.get("옵션입력")
        if pd.isna(option_str):
            continue

        try:
            option_dict, has_extra = parse_option_string(option_str)

            if not has_extra:
                all_expanded.append(row)
                continue

            grouped_sizes = group_sizes_by_extra_preserve_color(option_dict)

            color_match = re.search(r'색상\{(.+?)\}', option_str)
            color_str = color_match.group(1) if color_match else "기본"

            for group in grouped_sizes:
                sizes = group["sizes"]
                extra = group["extra_price"]

                new_option_str = f"색상{{{color_str}}}//사이즈{{{'|'.join(sizes)}}}"

                if pd.isna(row['공급가']) or not isinstance(row['공급가'], (int, float)):
                    continue

                supply_price = row['공급가'] + extra
                consumer_price = round(supply_price * 1.6)

                new_row = row.copy()
                new_row['옵션입력'] = new_option_str
                new_row['공급가'] = supply_price
                new_row['소비자가'] = consumer_price
                new_row['판매가'] = consumer_price
                new_row['상품명'] = update_product_name(new_row['상품명'], sizes, apply_update=True)

                all_expanded.append(new_row)

        except Exception as e:
            errors.append(f"Row {idx} 처리 중 오류 발생: {e}")
            continue

    return pd.DataFrame(all_expanded), errors


def clean_product_name(name):
    if not isinstance(name, str):
        return name
    if 'KC' in name:
        name = re.sub(r'^.*?KC\s*', '', name)
    idx = name.find('<')
    if idx != -1:
        name = name[:idx]
    return name.strip()


def extract_colors(option_str):
    if not isinstance(option_str, str):
        return []
    color_match = re.search(r'색상\{(.+?)\}', option_str)
    if not color_match:
        return []
    return [c.strip() for c in color_match.group(1).split('|')]


def extract_sizes_str(option_str):
    if not isinstance(option_str, str):
        return ""
    size_match = re.search(r'사이즈\{(.+?)\}', option_str)
    if not size_match:
        return ""
    return size_match.group(1)


def expand_by_color(df):
    all_rows = []
    for idx, row in df.iterrows():
        option_str = row.get("옵션입력")

        if pd.isna(option_str):
            new_row = row.copy()
            new_row['상품명'] = clean_product_name(new_row['상품명'])
            all_rows.append(new_row)
            continue

        colors = extract_colors(option_str)
        sizes_str = extract_sizes_str(option_str)

        if not colors:
            new_row = row.copy()
            new_row['상품명'] = clean_product_name(new_row['상품명'])
            all_rows.append(new_row)
            continue

        sizes_dash = sizes_str.replace('|', '-') if sizes_str else ""

        for color in colors:
            new_row = row.copy()
            new_row['상품명'] = clean_product_name(new_row['상품명'])
            new_row['옵션입력'] = color
            new_row['옵션 스타일'] = sizes_dash
            all_rows.append(new_row)

    return pd.DataFrame(all_rows)


def run_option_ver3(example_bytes, dated_filename):
    """예제.xlsx(bytes) -> (새 기린_df, 새 기린_bytes, 날짜파일_bytes, errors)"""
    df = pd.read_excel(io.BytesIO(example_bytes))
    kirin_df, errors = expand_rows(df)
    kirin_bytes_new = df_to_xlsx_bytes(kirin_df)

    expanded_color_df = expand_by_color(kirin_df)
    dated_bytes = df_to_xlsx_bytes(expanded_color_df)

    return kirin_df, kirin_bytes_new, dated_bytes, errors


# =========================================================
# 다음 라운드 준비: 두두사사 저장 / E,F,G 지운 다음 S파일 / 카테고리번호 갱신
# =========================================================

def make_next_sfile(kirin_bytes_new):
    """새 기린.xlsx 에서 E,F,G 값을 지운 사본 bytes 반환 (다음 라운드용 S파일)"""
    wb = load_wb(kirin_bytes_new, data_only=False)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        for col in ['E', 'F', 'G']:
            ws[f"{col}{r}"].value = None
    return wb_bytes(wb)


def extract_new_category_code(kirin_df):
    """새 기린_df의 E열 값에서 '|' 기준 맨 뒤 값을 추출.
    (예: 4337|2118|6136 -> 6136) 여러 행에 값이 있으면 첫 값을 사용하고,
    행마다 서로 다른 경우 경고를 반환.
    """
    if '상품분류 번호' not in kirin_df.columns:
        return None, ["'상품분류 번호'(E열) 컬럼을 찾을 수 없어 카테고리번호 갱신을 건너뜁니다."]

    warnings = []
    last_codes = []
    for val in kirin_df['상품분류 번호']:
        s = safe_str(val).strip()
        if not s:
            continue
        last = s.split('|')[-1].strip()
        if last:
            last_codes.append(last)

    if not last_codes:
        return None, ["E열(상품분류 번호)에 값이 없어 카테고리번호 갱신을 건너뜁니다."]

    new_code = last_codes[0]
    distinct = set(last_codes)
    if len(distinct) > 1:
        warnings.append(
            f"E열 맨 뒤 코드가 행마다 달랐습니다({sorted(distinct)}). "
            f"첫 번째 값인 '{new_code}'를 사용했습니다 — 값이 다르면 알려주세요."
        )

    return new_code, warnings


def update_category_file(category_bytes, new_code, target_row=29):
    """카테고리번호.xlsx 의 target_row 행, A열('원본')을 new_code로 갱신"""
    wb = load_wb(category_bytes, data_only=False)
    ws = wb.active

    old_value = ws.cell(row=target_row, column=1).value

    value_to_set = new_code
    try:
        value_to_set = int(new_code)
    except (TypeError, ValueError):
        pass

    ws.cell(row=target_row, column=1).value = value_to_set

    return wb_bytes(wb), old_value, value_to_set


# =========================================================
# 2단계: 자동업데이트파일생성(2026).py 로직
# =========================================================

def process_cafe24(dudu_bytes, today_str):
    df = pd.read_excel(io.BytesIO(dudu_bytes), dtype=str)
    column_name = '상품명'

    if column_name not in df.columns:
        raise PipelineError(f"두두사사 데이터에 '{column_name}' 컬럼이 없습니다.")

    outputs = {}

    # 2번시트 그룹
    df_sheet2 = df.copy()

    # 김하늘: 소비자가(T열)/판매가(V열) = 공급가(U열) * 1.7
    df_kim_haneul = df_sheet2.copy()
    if all(c in df_kim_haneul.columns for c in ['소비자가', '공급가', '판매가']):
        supply_numeric = pd.to_numeric(df_kim_haneul['공급가'], errors='coerce')
        computed_price = (supply_numeric * 1.7).round()
        df_kim_haneul['소비자가'] = computed_price
        df_kim_haneul['판매가'] = computed_price
    outputs['김하늘.xlsx'] = df_to_xlsx_bytes(df_kim_haneul, sheet_name="2번시트")

    # 서송이: 적립금(AB열) 값을 모두 비움
    df_seo = df_sheet2.copy()
    if '적립금' in df_seo.columns:
        df_seo['적립금'] = None
    outputs['서송이.xlsx'] = df_to_xlsx_bytes(df_seo, sheet_name="2번시트")

    for name in ['이지민', '해피', '김다혜']:
        outputs[f'{name}.xlsx'] = df_to_xlsx_bytes(df_sheet2, sheet_name="2번시트")

    # 3번시트 그룹
    df_sheet3 = df.copy()
    df_sheet3[column_name] = (
        df_sheet3[column_name].astype(str)
        .str.replace(r'.*KC', '', regex=True)
        .str.replace(r'<.*', '', regex=True)
        .str.replace(' ', '', regex=False)
    )
    for name in ['박미리']:
        outputs[f'{name}.xlsx'] = df_to_xlsx_bytes(df_sheet3, sheet_name="3번시트")

    # 5번시트 그룹
    df_sheet5 = df.copy()
    df_sheet5[column_name] = (
        df_sheet5[column_name].astype(str)
        .str.replace(r'<.*', '', regex=True)
        .str.replace('KC', '', regex=False)
    )
    for name in ['오토', '오마베', '최민기', '현건주', '목윤희', '최은경', '남부센']:
        outputs[f'{name}.xlsx'] = df_to_xlsx_bytes(df_sheet5, sheet_name="5번시트")

    # 백종열.xlsx
    df_baek = df_sheet5.copy()
    o_column = '상품 상세설명'
    if o_column in df_baek.columns:
        df_baek[o_column] = (
            '<p align=center><img src="https://mododome.diskn.com/p7Mch6ng1C" /><br><br></p>'
            + df_baek[o_column].astype(str)
            + '<p align=center><img src="https://mododome.diskn.com/r7Mch6n6Ee" /><br><br></p>'
        )
    outputs['백종열.xlsx'] = df_to_xlsx_bytes(df_baek, sheet_name="5번시트")

    # X파일
    x_filename = f'X{today_str}.xlsx'
    outputs[x_filename] = df_to_xlsx_bytes(df_sheet5, sheet_name="5번시트")

    # 백성희.xlsx
    df_modified = df.copy()
    if '진열상태' in df_modified.columns:
        df_modified['진열상태'] = df_modified['진열상태'].replace('Y', 'Y')
    if '판매상태' in df_modified.columns:
        df_modified['판매상태'] = df_modified['판매상태'].replace('Y', 'Y')
    df_modified[column_name] = (
        df_modified[column_name].astype(str)
        .str.replace(r'<.*', '', regex=True)
        .str.replace('KC', '', regex=False)
    )
    outputs['백성희.xlsx'] = df_to_xlsx_bytes(df_modified, sheet_name="5번시트")

    # 모나마켓: 진열상태(C열)/판매상태(D열)을 N으로 생성
    for label in ['모나마켓']:
        df_choi = df_modified.copy()
        if '진열상태' in df_choi.columns:
            df_choi['진열상태'] = 'N'
        if '판매상태' in df_choi.columns:
            df_choi['판매상태'] = 'N'
        df_choi[column_name] = df_choi[column_name].astype(str) + f' 26가을 {label}'
        outputs[f'{label}.xlsx'] = df_to_xlsx_bytes(df_choi, sheet_name="5번시트")

    # 문정희.xlsx는 1단계(build_example_state)에서 이미 최종본으로 생성되므로
    # 2단계에서는 별도로 만들지 않습니다 (중복/덮어쓰기 방지).

    # 이모이모.xlsx
    df_sheet6 = df.copy()
    df_sheet6[column_name] = (
        df_sheet6[column_name].astype(str)
        .str.replace(r'<.*', '', regex=True)
        .str.replace('KC', '', regex=False)
    )
    outputs['이모이모.xlsx'] = df_to_xlsx_bytes(df_sheet6, sheet_name="6번시트")

    return outputs, df, x_filename


def parse_options(option_str):
    text = safe_str(option_str).strip()
    if not text:
        return '', '', '', ''

    try:
        color_match = re.search(r'(색상|컬러)\{([^\}]*)\}', text)
        size_match = re.search(r'사이즈\{([^\}]*)\}', text)

        colors = []
        sizes = []

        if color_match:
            colors = [x.strip() for x in color_match.group(2).split('|') if x.strip()]
        if size_match:
            sizes = [x.strip() for x in size_match.group(1).split('|') if x.strip()]

        if not colors:
            color_match2 = re.search(r'(색상|컬러)\s*[:：]\s*([^/\n]+)', text)
            if color_match2:
                colors = [x.strip() for x in color_match2.group(2).split('|') if x.strip()]

        if not sizes:
            size_match2 = re.search(r'사이즈\s*[:：]\s*([^/\n]+)', text)
            if size_match2:
                sizes = [x.strip() for x in size_match2.group(1).split('|') if x.strip()]

        if not colors and not sizes:
            cleaned = re.sub(r'<.*?>', '', text).strip()
            if cleaned:
                return '옵션', cleaned, '0', '50'
            return '', '', '', ''

        option_name_parts = []
        if colors:
            option_name_parts.append('컬러')
        if sizes:
            option_name_parts.append('사이즈')
        option_name = '\n'.join(option_name_parts)

        value_parts = []
        if colors:
            value_parts.append(','.join(colors))
        if sizes:
            value_parts.append(','.join(sizes))
        option_value = '\n'.join(value_parts)

        price_parts = []
        if colors:
            price_parts.append(','.join(['0'] * len(colors)))
        if sizes:
            price_parts.append(','.join(['0'] * len(sizes)))
        option_price = '\n'.join(price_parts)

        option_stock = ','.join(['50'] * len(colors)) if colors else ''

        return option_name, option_value, option_price, option_stock

    except Exception:
        return '', '', '', ''


REQUIRED_SMARTSTORE_HEADERS = [
    '판매자 상품코드', '카테고리코드', '상품명', '판매가', '단위가격 사용여부',
    '관부가세', '옵션형태', '옵션명', '옵션값', '옵션가', '옵션 재고수량',
    '대표이미지', '상세설명',
]

SMARTSTORE_SRC_MAP = {
    '판매자 상품코드': 'U',
    '상품명': 'H',
    '판매가': 'V',
    '대표이미지': 'AT',
    '상세설명': 'O',
    '브랜드': 'L',
    '제조사': 'L',
    '제조일자': 'BB',
    '옵션원본': 'AJ',
}

SMARTSTORE_FALLBACK_DEFAULTS_RAW = {
    '상품상태': '신상품',
    '단위가격 사용여부': 'N',
    '부가세': '과세상품',
    '관부가세': '부과 대상 아님',
    '재고수량': 100,
    '옵션형태': '조합형',
    '원산지코드': '03',
    '복수원산지여부': 'N',
    '미성년자 구매': 'Y',
    '배송비 템플릿코드': 3297037,
    '반품배송비': 2500,
    '교환배송비': 5000,
    '별도설치비': 'N',
    'A/S 템플릿코드': 3248096,
}

SMARTSTORE_BLANK_HEADERS = [
    '표시용량', '표시단위', '총용량',
    '사이즈\n상품군', '사이즈\n사이즈명', '사이즈\n상세 사이즈', '사이즈\n모델명',
]


def _build_smartstore_rows(ws_src, row_numbers, template_bytes):
    """ws_src의 지정된 행 번호(row_numbers, 1-based, ws_src 기준)들만 골라
    스마트스토어 템플릿 한 장에 채워 넣은 워크북 bytes를 반환."""
    wb_template = load_wb(template_bytes, data_only=False)
    ws_template = wb_template.active

    template_headers = build_header_map(ws_template, header_row=2)

    missing_headers = [h for h in REQUIRED_SMARTSTORE_HEADERS if normalize_header(h) not in template_headers]
    if missing_headers:
        raise PipelineError(f"스마트스토어 템플릿에서 아래 헤더를 찾을 수 없습니다: {missing_headers}")

    template_default_values = {}
    for norm_header, col_idx in template_headers.items():
        value = ws_template.cell(row=3, column=col_idx).value
        if value is not None:
            template_default_values[norm_header] = value

    fallback_defaults = {normalize_header(k): v for k, v in SMARTSTORE_FALLBACK_DEFAULTS_RAW.items()}
    final_defaults = fallback_defaults.copy()
    final_defaults.update(template_default_values)

    current_row = 3
    for i in row_numbers:
        for col in range(1, ws_template.max_column + 1):
            ws_template.cell(row=current_row, column=col).value = None

        for norm_header, value in final_defaults.items():
            if norm_header in template_headers:
                ws_template.cell(row=current_row, column=template_headers[norm_header]).value = value

        set_cell_by_header(ws_template, current_row, template_headers, '단위가격 사용여부', 'N')
        set_cell_by_header(ws_template, current_row, template_headers, '관부가세', '부과 대상 아님')

        for h in SMARTSTORE_BLANK_HEADERS:
            set_cell_by_header(ws_template, current_row, template_headers, h, None)

        for target_header, src_col_letter in SMARTSTORE_SRC_MAP.items():
            if target_header == '옵션원본':
                continue
            src_value = ws_src.cell(row=i, column=get_col_index(src_col_letter)).value
            set_cell_by_header(ws_template, current_row, template_headers, target_header, src_value)

        option_raw = ws_src.cell(row=i, column=get_col_index(SMARTSTORE_SRC_MAP['옵션원본'])).value
        option_name, option_value, option_price, option_stock = parse_options(option_raw)

        if option_name or option_value:
            set_cell_by_header(ws_template, current_row, template_headers, '옵션형태', '조합형')
            set_cell_by_header(ws_template, current_row, template_headers, '옵션명', option_name, wrap_text=True)
            set_cell_by_header(ws_template, current_row, template_headers, '옵션값', option_value, wrap_text=True)
            set_cell_by_header(ws_template, current_row, template_headers, '옵션가', option_price, wrap_text=True)
            set_cell_by_header(ws_template, current_row, template_headers, '옵션 재고수량', option_stock, wrap_text=True)
        else:
            set_cell_by_header(ws_template, current_row, template_headers, '옵션형태', '단일상품')
            set_cell_by_header(ws_template, current_row, template_headers, '옵션명', '')
            set_cell_by_header(ws_template, current_row, template_headers, '옵션값', '')
            set_cell_by_header(ws_template, current_row, template_headers, '옵션가', '')
            set_cell_by_header(ws_template, current_row, template_headers, '옵션 재고수량', '')

        current_row += 1

    return wb_bytes(wb_template)


def convert_to_smartstore_chunks(x_bytes, template_bytes, chunk_size=500):
    """X파일 -> 스마트스토어 템플릿 변환. 실제 데이터 행 수가 chunk_size(기본 500)를
    초과하면 앞에서부터 chunk_size행씩 잘라 여러 개의 워크북으로 나눠 만듭니다.

    반환: [{'start': 0-based 시작 인덱스, 'end': 끝(미포함) 인덱스, 'row_count': 행수, 'bytes': ...}, ...]
    (start/end는 '유효 데이터 행'(상품명이 있는 행) 목록 기준 0-based 인덱스이며,
    두두사사 DataFrame(df_dudu)의 행 순서와 1:1로 대응합니다.)
    """
    wb_src = load_wb(x_bytes, data_only=True)
    ws_src = wb_src.active

    valid_rows = [i for i in range(2, ws_src.max_row + 1) if ws_src.cell(row=i, column=get_col_index('H')).value]

    if not valid_rows:
        return []

    chunks = []
    for start in range(0, len(valid_rows), chunk_size):
        row_numbers = valid_rows[start:start + chunk_size]
        chunk_bytes = _build_smartstore_rows(ws_src, row_numbers, template_bytes)
        chunks.append({
            'start': start,
            'end': start + len(row_numbers),
            'row_count': len(row_numbers),
            'bytes': chunk_bytes,
        })

    return chunks


def apply_category_mapping(df_dudu, outputs, n_file_info, mapping_df, target_col='스마트오토'):
    """마켓별 파일들의 E/F/G열, 그리고 N파일(들)의 B열(카테고리코드)에 카테고리 매핑을 적용.

    n_file_info: [{'filename': ..., 'start': 0-based, 'end': 0-based(미포함)}, ...]
    (N파일이 500행 단위로 여러 개로 쪼개져 있을 수 있으므로 각 청크의 start/end로
    df_dudu의 올바른 행 구간을 찾아 매핑합니다.)
    """
    warnings = []
    required_df_cols = ['상품분류 번호', '상품분류 신상품영역', '상품분류 추천상품영역']
    for col in required_df_cols:
        if col not in df_dudu.columns:
            raise PipelineError(f"두두사사 데이터에 '{col}' 컬럼이 없습니다.")
    if '원본' not in mapping_df.columns:
        raise PipelineError("카테고리번호.xlsx에 '원본' 컬럼이 없습니다.")

    for filename in list(outputs.keys()):
        if not filename.endswith('.xlsx'):
            continue
        stem = filename[:-5]
        if stem not in mapping_df.columns:
            continue

        mapping_dict = dict(zip(str_series(mapping_df['원본']), str_series(mapping_df[stem])))

        wb = load_wb(outputs[filename], data_only=False)
        ws = wb.active

        headers = {safe_str(cell.value).strip(): cell.column_letter for cell in ws[1] if cell.value is not None}
        needed_headers = ['상품분류 번호', '상품분류 신상품영역', '상품분류 추천상품영역']
        if not all(k in headers for k in needed_headers):
            warnings.append(f"{stem}.xlsx -> 필요한 컬럼이 없어 카테고리 매핑을 건너뜀")
            continue

        for i in range(2, len(df_dudu) + 2):
            raw_e = safe_str(df_dudu.loc[i - 2, '상품분류 번호'])
            raw_f = safe_str(df_dudu.loc[i - 2, '상품분류 신상품영역'])
            raw_g = safe_str(df_dudu.loc[i - 2, '상품분류 추천상품영역'])

            mapped = [mapping_dict.get(code.strip(), code.strip()) for code in raw_e.split('|') if code.strip()]
            ws[f"{headers['상품분류 번호']}{i}"] = '|'.join(mapped)
            ws[f"{headers['상품분류 신상품영역']}{i}"] = raw_f
            ws[f"{headers['상품분류 추천상품영역']}{i}"] = raw_g

        outputs[filename] = wb_bytes(wb)

    if n_file_info:
        if target_col not in mapping_df.columns:
            warnings.append(f"'{target_col}' 컬럼이 카테고리번호.xlsx에 없어 스마트스토어(N) 파일 매핑을 생략했습니다.")
        else:
            t_map = dict(zip(str_series(mapping_df['원본']), str_series(mapping_df[target_col])))
            for info in n_file_info:
                fname = info['filename']
                if fname not in outputs:
                    continue
                wb_n = load_wb(outputs[fname], data_only=False)
                ws_n = wb_n.active
                chunk_len = info['end'] - info['start']
                for local_row in range(chunk_len):
                    df_idx = info['start'] + local_row
                    raw_e = safe_str(df_dudu.loc[df_idx, '상품분류 번호'])
                    first_code = raw_e.split('|')[0].strip() if raw_e else ''
                    mapped_value = t_map.get(first_code, '')
                    ws_n[f'B{local_row + 3}'] = mapped_value
                outputs[fname] = wb_bytes(wb_n)

    return outputs, warnings


def apply_category_mapping_to_self(stage1_outputs, mapping_df, keys=('문정희.xlsx', '곽충현.xlsx')):
    """1단계 산출물인 문정희.xlsx / 곽충현.xlsx의 E열('상품분류 번호')을,
    2단계 판매처 파일들과 동일한 방식(카테고리번호.xlsx의 '원본' -> 해당 판매처 컬럼)으로
    매핑해 갱신합니다.

    2단계 apply_category_mapping()과 달리 이 두 파일은 0단계(옵션ver3 확장 전)에서
    만들어져 두두사사(df_dudu)와 행 수/순서가 다를 수 있으므로, df_dudu 값을 위치로
    가져오는 대신 파일 자신이 이미 가지고 있는 E열 값을 그대로 매핑 키로 사용합니다.
    F/G열은 2단계와 동일하게(원본 코드 그대로) 건드리지 않습니다.
    """
    warnings = []
    if mapping_df is None or '원본' not in mapping_df.columns:
        return stage1_outputs, warnings

    for filename in keys:
        if filename not in stage1_outputs:
            continue
        stem = filename[:-5]
        if stem not in mapping_df.columns:
            warnings.append(f"카테고리번호.xlsx에 '{stem}' 컬럼이 없어 {filename}의 E열 카테고리 매핑을 건너뜀")
            continue

        mapping_dict = dict(zip(str_series(mapping_df['원본']), str_series(mapping_df[stem])))

        wb = load_wb(stage1_outputs[filename], data_only=False)
        ws = wb.active

        e_col = None
        for cell in ws[1]:
            if safe_str(cell.value).strip() == '상품분류 번호':
                e_col = cell.column
                break
        if e_col is None:
            warnings.append(f"{filename}에서 '상품분류 번호'(E열) 컬럼을 찾을 수 없어 카테고리 매핑을 건너뜀")
            continue

        for r in range(2, ws.max_row + 1):
            raw_e = safe_str(ws.cell(row=r, column=e_col).value)
            if not raw_e:
                continue
            mapped = [mapping_dict.get(code.strip(), code.strip()) for code in raw_e.split('|') if code.strip()]
            ws.cell(row=r, column=e_col).value = '|'.join(mapped)

        stage1_outputs[filename] = wb_bytes(wb)

    return stage1_outputs, warnings


def append_mapping_to_auto_file(outputs, n_file_info):
    """N파일(들)의 B열(카테고리코드) 값을, N파일이 여러 청크로 나뉘어 있어도
    전체 행 순서 그대로 이어붙여 오토.xlsx 상품명 뒤에 추가."""
    warnings = []
    if not n_file_info or '오토.xlsx' not in outputs:
        warnings.append("N파일 또는 오토.xlsx가 없어 오토 파일 상품명 매핑 추가를 건너뜀")
        return outputs, warnings

    mappings = []
    for info in n_file_info:
        fname = info['filename']
        if fname not in outputs:
            continue
        wb_n = load_wb(outputs[fname], data_only=True)
        ws_n = wb_n.active
        chunk_len = info['end'] - info['start']
        for local_row in range(chunk_len):
            value = ws_n[f'B{local_row + 3}'].value
            mappings.append(str(value) if value is not None else '')

    wb_auto = load_wb(outputs['오토.xlsx'], data_only=False)
    ws_auto = wb_auto.active

    h_col_letter = None
    for cell in ws_auto[1]:
        if cell.value == '상품명':
            h_col_letter = cell.column_letter
            break

    if not h_col_letter:
        warnings.append("오토.xlsx에서 '상품명' 컬럼을 찾을 수 없어 매핑 추가를 건너뜀")
        return outputs, warnings

    for idx, map_val in enumerate(mappings):
        row_num = idx + 2
        cell = ws_auto[f"{h_col_letter}{row_num}"]
        original = str(cell.value) if cell.value else ''
        cell.value = f"{original}*{map_val}"

    outputs['오토.xlsx'] = wb_bytes(wb_auto)
    return outputs, warnings


def clear_columns(outputs, x_filename):
    targets = ['금.xlsx', '오토.xlsx', '0.xlsx', x_filename]
    for fname in targets:
        if fname in outputs:
            wb = load_wb(outputs[fname], data_only=False)
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                for col in ['E', 'F', 'G']:
                    ws[f"{col}{row}"].value = None
            outputs[fname] = wb_bytes(wb)
    return outputs


# =========================================================
# 전체 오케스트레이션
# =========================================================

def run_pipeline(
    kirin_bytes,
    sfile_bytes,
    category_bytes,
    template_bytes=None,
    dated_filename_date=None,
    next_sfile_date=None,
    category_row=29,
    smartstore_target_col='스마트오토',
):
    """전체 파이프라인 실행.

    반환: {
        'stage1': {파일명: bytes, ...},   # 예제/새기린/날짜파일/두두사사/다음S파일/카테고리번호/곽충현/문정희
                                          # (곽충현/문정희는 카테고리번호.xlsx의 '곽충현'/'문정희' 컬럼이
                                          #  있으면 그 값으로 E열이 매핑된 상태로 생성됩니다)
        'stage2': {파일명: bytes, ...},   # 18개 판매처 파일(문정희 제외, 1단계에서 생성) + N파일(500행 초과 시 -1, -2...로 분할)
        'logs': [str, ...],
        'errors': [str, ...],
    }
    """
    logs = []
    errors = []

    today_str = datetime.today().strftime('%Y%m%d')
    dated_str = dated_filename_date or today_str
    next_s_str = next_sfile_date or today_str

    # 0단계
    state = build_example_state(kirin_bytes, sfile_bytes)
    logs += state['warnings']

    # 1단계 (옵션ver3)
    kirin_df, kirin_bytes_new, dated_bytes, opt_errors = run_option_ver3(
        state['example_bytes'], dated_str
    )
    errors += opt_errors

    dudu_bytes = kirin_bytes_new  # 두두사사.xlsx = 새로 생성된 기린.xlsx와 동일 내용
    next_sfile_bytes = make_next_sfile(kirin_bytes_new)

    # 카테고리번호 갱신
    new_code, code_warnings = extract_new_category_code(kirin_df)
    logs += code_warnings
    if new_code is not None:
        category_bytes_updated, old_val, new_val = update_category_file(
            category_bytes, new_code, target_row=category_row
        )
        logs.append(f"카테고리번호.xlsx {category_row}행 A열: '{old_val}' -> '{new_val}' 로 갱신했습니다.")
    else:
        category_bytes_updated = category_bytes

    stage1 = {
        '예제.xlsx': state['example_bytes'],
        '기린.xlsx': kirin_bytes_new,
        f'{dated_str}.xlsx': dated_bytes,
        '두두사사.xlsx': dudu_bytes,
        f'S{next_s_str}.xlsx': next_sfile_bytes,
        '카테고리번호.xlsx': category_bytes_updated,
        '곽충현.xlsx': state['kwak_bytes'],
        '문정희.xlsx': state['moon_bytes'],
    }

    # 2단계 (자동업데이트파일생성 2026)
    try:
        mapping_df = pd.read_excel(io.BytesIO(category_bytes_updated), dtype=str)
    except Exception as e:
        errors.append(f"카테고리번호.xlsx를 읽는 중 오류: {e}")
        mapping_df = None

    if mapping_df is not None:
        stage1, moon_kwak_warnings = apply_category_mapping_to_self(stage1, mapping_df)
        logs += moon_kwak_warnings

    stage2 = {}
    try:
        outputs, df_dudu, x_filename = process_cafe24(dudu_bytes, today_str)
        stage2.update(outputs)

        n_file_info = []
        if template_bytes is not None:
            try:
                chunks = convert_to_smartstore_chunks(stage2[x_filename], template_bytes, chunk_size=500)
                if len(chunks) == 1:
                    fname = f'N{today_str}.xlsx'
                    stage2[fname] = chunks[0]['bytes']
                    n_file_info.append({'filename': fname, 'start': chunks[0]['start'], 'end': chunks[0]['end']})
                elif len(chunks) > 1:
                    for idx, c in enumerate(chunks, start=1):
                        fname = f'N{today_str}-{idx}.xlsx'
                        stage2[fname] = c['bytes']
                        n_file_info.append({'filename': fname, 'start': c['start'], 'end': c['end']})
                    total_rows = sum(c['row_count'] for c in chunks)
                    logs.append(
                        f"N파일 데이터가 총 {total_rows}행으로 500행을 초과해 "
                        f"{len(chunks)}개 파일로 나눠 생성했습니다: {', '.join(i['filename'] for i in n_file_info)}"
                    )
            except Exception as e:
                errors.append(f"스마트스토어(N파일) 변환 중 오류: {e}")
        else:
            logs.append("스마트스토어기본.xlsx 템플릿이 제공되지 않아 N파일(스마트스토어 변환) 생성을 건너뛰었습니다.")

        if mapping_df is not None:
            stage2, map_warnings = apply_category_mapping(
                df_dudu, stage2, n_file_info, mapping_df, target_col=smartstore_target_col
            )
            logs += map_warnings

            stage2, auto_warnings = append_mapping_to_auto_file(stage2, n_file_info)
            logs += auto_warnings

        stage2 = clear_columns(stage2, x_filename)

    except PipelineError as e:
        errors.append(str(e))
    except Exception as e:
        errors.append(f"2단계 처리 중 예상치 못한 오류: {e}")

    return {
        'stage1': stage1,
        'stage2': stage2,
        'logs': logs,
        'errors': errors,
    }
